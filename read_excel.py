import openpyxl
import pandas as pd

# Load the workbook
wb = openpyxl.load_workbook('Max.xlsx')
print(f"Sheets: {wb.sheetnames}\n")

# Read each sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print('='*60)
    ws = wb[sheet_name]
    
    # Print first 30 rows
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(cell is not None for cell in row):  # Skip completely empty rows
            print(f"Row {i}: {row}")
        if i >= 30:
            break

print("\n\nNow using pandas for better formatting:")
print("="*60)

# Try reading with pandas
for sheet_name in wb.sheetnames:
    print(f"\n\nSheet: {sheet_name}")
    print("-"*60)
    df = pd.read_excel('Max.xlsx', sheet_name=sheet_name)
    print(df.head(20))
    print(f"\nColumns: {df.columns.tolist()}")
