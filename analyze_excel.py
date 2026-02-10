import openpyxl

wb = openpyxl.load_workbook('Max.xlsx', data_only=True)
ws = wb.active

print("LOTTERY INVESTMENT CALCULATOR ANALYSIS")
print("="*70)
print("\nINPUTS:")
print("-"*70)
print(f"Lottery Amount: ${ws['D3'].value:,.0f}")
print(f"Cash Sum (52%): ${ws['D4'].value:,.2f}")
print(f"Federal Taxes (37%): ${ws['D5'].value:,.2f}")
print(f"State Taxes (5.5%): ${ws['D6'].value:,.2f}")
print(f"Net Amount: ${ws['D7'].value:,.2f}")
print(f"Canadian Conversion (1.35x): ${ws['D8'].value:,.2f}")
print(f"\nInvest 90%: ${ws['D10'].value:,.2f}")
print(f"Fun Money (10%): ${ws['G7'].value:,.2f}")
print(f"Net Daily Income: ${ws['G8'].value:,.2f}")

print("\n\nINVESTMENT CYCLES (90-day periods at 4.5% annual return):")
print("-"*70)
print(f"{'Cycle':<10} {'Principal Start':<18} {'Interest':<15} {'Taxes':<15} {'Withdrawal':<15} {'Total Out':<15} {'Reinvest':<15} {'Principal End':<18}")
print("-"*70)

for row in range(13, 21):  # Cycles 1-8
    cycle = ws[f'C{row}'].value
    principal_start = ws[f'D{row}'].value
    interest = ws[f'E{row}'].value
    taxes = ws[f'F{row}'].value
    withdrawal = ws[f'G{row}'].value
    total_out = ws[f'H{row}'].value
    reinvest = ws[f'I{row}'].value
    principal_end = ws[f'J{row}'].value
    
    print(f"{cycle:<10} ${principal_start:>15,.0f} ${interest:>13,.0f} ${taxes:>13,.0f} ${withdrawal:>13,.0f} ${total_out:>13,.0f} ${reinvest:>13,.0f} ${principal_end:>15,.0f}")

print("\n\nKEY FORMULAS:")
print("-"*70)
print("• Interest per cycle: Principal × 4.5% × (90/365)")
print("• Taxes on interest: Interest × 53.53%")
print("• Personal withdrawal: Interest × 10%")
print("• Total withdrawal: Taxes + Personal withdrawal")
print("• Reinvestment: Interest - Total withdrawal")
print("• Next cycle principal: Previous principal + Reinvestment")

total_withdrawn = ws['G7'].value
print(f"\n\nTOTAL PERSONAL WITHDRAWALS (8 cycles): ${total_withdrawn:,.2f}")
print(f"Average per cycle: ${total_withdrawn/8:,.2f}")
print(f"Daily income: ${ws['G8'].value:,.2f}")
print(f"Annual income: ${ws['G8'].value * 365:,.2f}")
